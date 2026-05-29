"""Generate notebooks/01_phase1_cme_reference_demo.ipynb.

This builder keeps the notebook authoring reproducible. The notebook itself
calls only reusable services under src/large_pdf_extractor/ — no pipeline logic
is implemented inline.
"""

from __future__ import annotations

from pathlib import Path

import nbformat as nbf

nb = nbf.v4.new_notebook()
cells: list = []


def md(text: str) -> None:
    cells.append(nbf.v4.new_markdown_cell(text.strip("\n")))


def code(text: str) -> None:
    cells.append(nbf.v4.new_code_cell(text.strip("\n")))


md(
    """
# Phase 1 — Large PDF Extraction: CME Reference Demo

This notebook demonstrates the **dictionary-first** large-PDF extraction
pipeline end to end on the reference document
`data/input/Metals_Option_Products.pdf`.

It uses **only** the reusable services under `src/large_pdf_extractor/` and runs
with the deterministic **FakeLLMProvider** — no API keys or paid calls required.

The CME bulletin is a *reference example* of a difficult large PDF (repeated
headers/footers, dense tables, product sections, disclaimers). The pipeline
itself is generic and Finance-Market-Ops aware via configurable hints only.
"""
)

md("## 1. Environment and path setup")
code(
    """
import sys, os, json
from pathlib import Path

# Make the src/ package importable when running from the notebooks/ folder.
ROOT = Path.cwd()
if (ROOT / "src").exists():
    PROJECT_ROOT = ROOT
elif (ROOT.parent / "src").exists():
    PROJECT_ROOT = ROOT.parent
else:
    raise RuntimeError("Could not locate project root containing src/.")

SRC = PROJECT_ROOT / "src"
if str(SRC) not in sys.path:
    sys.path.insert(0, str(SRC))

print("Project root:", PROJECT_ROOT)
import large_pdf_extractor
print("large_pdf_extractor version:", large_pdf_extractor.__version__)
"""
)

md("## 2. Validate that the reference input files exist")
code(
    """
PDF_PATH = PROJECT_ROOT / "data" / "input" / "Metals_Option_Products.pdf"
RTF_PATH = PROJECT_ROOT / "data" / "input" / "metal_options_summary.rtf"
OUTPUT_DIR = PROJECT_ROOT / "data" / "output"

pdf_exists = PDF_PATH.exists()
rtf_exists = RTF_PATH.exists()
print(f"PDF present:  {pdf_exists}  ({PDF_PATH})")
print(f"RTF present:  {rtf_exists}  ({RTF_PATH})")

if not pdf_exists:
    print(
        "\\nNOTE: The reference PDF is missing. Place it at the path above to run "
        "the full demo. The remaining cells require it."
    )
"""
)

md("## 3. Inspect the PDF profile (PyMuPDF baseline)")
code(
    """
from large_pdf_extractor.parsing import get_parser
from large_pdf_extractor.domain.models import ParseStrategy
from large_pdf_extractor.profiling.document_profiler import (
    DocumentProfiler, select_representative_chunks,
)
from large_pdf_extractor.chunking.chunker import ChunkingService

assert pdf_exists, "Reference PDF required for this cell."

parser = get_parser(ParseStrategy.PYMUPDF)
parsed = parser.parse(str(PDF_PATH))
profile = DocumentProfiler().profile(parsed)

print("Document id:           ", parsed.document_id)
print("Page count:            ", profile.page_count)
print("Total characters:      ", profile.total_char_count)
print("Low-text pages:        ", len(profile.empty_or_low_text_pages))
print("Detected table pages:  ", profile.detected_table_count)
print("Repeated headers:      ", profile.repeated_header_candidates[:3])
print("Repeated footers:      ", profile.repeated_footer_candidates[:3])
print("Sample headings:       ", profile.detected_headings[:5])
"""
)

md("## 4. Dictionary proposal — BEFORE any extraction (dictionary-first)")
code(
    """
from large_pdf_extractor.dictionary.generator import DictionaryService

chunks = ChunkingService().chunk(parsed, profile)
representative = select_representative_chunks(profile, chunks)

dictionary = DictionaryService().generate_proposed_dictionary(profile, representative)
print("Proposed dictionary:", dictionary.dictionary_id)
print("Item count:", len(dictionary.items))
for item in dictionary.items:
    print(f"  - {item.item_id:32s} [{item.expected_type.value}] section={item.document_section}")
"""
)

md("## 5. Validate the dictionary with Pydantic")
code(
    """
from large_pdf_extractor.dictionary.loader import DictionaryLoader
from large_pdf_extractor.domain.models import ExtractionDictionary

validated = DictionaryLoader().validate(dictionary.model_dump())
assert isinstance(validated, ExtractionDictionary)
assert validated.items, "Dictionary must contain items"
print("Dictionary validated OK with", len(validated.items), "items.")
"""
)

md(
    """
## 6. PyMuPDF path — parse / profile / chunk / extract / render

We run the full pipeline through the orchestrator (the same code the CLI uses).
"""
)
code(
    """
from large_pdf_extractor.app.config import build_run_config
from large_pdf_extractor.app.pipeline import Phase1Pipeline

py_config = build_run_config(
    pdf_path=str(PDF_PATH),
    output_dir=str(OUTPUT_DIR),
    strategy=ParseStrategy.PYMUPDF,
    llm_provider="fake",
    max_chunks=12,
)
py_state = Phase1Pipeline(py_config).run()
py_run_dir = OUTPUT_DIR / py_state.document_id / py_state.run_id
print("PyMuPDF run id:", py_state.run_id)
print("Artifacts written:", len(py_state.artifacts))
print("Run dir:", py_run_dir)
"""
)

md("## 7. Docling path — runs if installed, otherwise graceful skip")
code(
    """
from large_pdf_extractor.parsing.docling_parser import DoclingParser

# Docling runs PyTorch models. On Apple Silicon (macOS/MPS) this can crash mid-
# inference, so the parser defaults to the CPU accelerator. Override with the
# LPE_DOCLING_DEVICE env var (cpu | mps | cuda | auto) before importing if needed.
parser_dl = DoclingParser()
docling_available = parser_dl.is_available()
print("Docling available:", docling_available)
print("Docling device:   ", parser_dl.device)

dl_config = build_run_config(
    pdf_path=str(PDF_PATH),
    output_dir=str(OUTPUT_DIR),
    strategy=ParseStrategy.DOCLING,
    llm_provider="fake",
    max_chunks=12,
)
dl_state = Phase1Pipeline(dl_config).run()
dl_run_dir = OUTPUT_DIR / dl_state.document_id / dl_state.run_id

dl_parsed_meta = json.loads((dl_run_dir / "parsed_document.docling.json").read_text())["metadata"]
if dl_parsed_meta.get("skipped"):
    # A skip is never silent/blank: the reason is always surfaced here.
    print("Docling path was skipped. Reason(s):")
    for w in dl_parsed_meta.get("warnings", []):
        print("   -", w)
else:
    print(f"Docling parsed {dl_state} via device={dl_parsed_meta.get('device')}")
    print("   pages:", json.loads((dl_run_dir / 'parsed_document.docling.json').read_text())['page_count'])
    print("   docling tables:", dl_parsed_meta.get("docling_table_count"))
"""
)

md("## 8. Compare mode — one shared dictionary across both parser paths")
code(
    """
cmp_config = build_run_config(
    pdf_path=str(PDF_PATH),
    output_dir=str(OUTPUT_DIR),
    strategy=ParseStrategy.COMPARE,
    llm_provider="fake",
    max_chunks=12,
)
cmp_state = Phase1Pipeline(cmp_config).run()
cmp_run_dir = OUTPUT_DIR / cmp_state.document_id / cmp_state.run_id
print("Compare run id:", cmp_state.run_id)
print("Run dir:", cmp_run_dir)
print("Comparison report present:",
      (cmp_run_dir / "comparison_report.json").exists())
"""
)

md("## 9. Generated JSON and Markdown artifacts")
code(
    """
print("PyMuPDF run artifacts:")
for p in sorted(py_run_dir.iterdir()):
    print("  ", p.name)
print("\\nCompare run artifacts:")
for p in sorted(cmp_run_dir.iterdir()):
    print("  ", p.name)
"""
)

md("## 10. Reload and validate generated outputs through Pydantic")
code(
    """
from large_pdf_extractor.domain.models import (
    ExtractionResult, ComparisonReport, ParsedDocument, DocumentProfile,
)

py_result = ExtractionResult.model_validate_json(
    (py_run_dir / "extraction_result.pymupdf.json").read_text()
)
cmp_report = ComparisonReport.model_validate_json(
    (cmp_run_dir / "comparison_report.json").read_text()
)
reloaded_doc = ParsedDocument.model_validate_json(
    (py_run_dir / "parsed_document.pymupdf.json").read_text()
)
reloaded_profile = DocumentProfile.model_validate_json(
    (py_run_dir / "document_profile.pymupdf.json").read_text()
)

populated = sum(1 for v in py_result.values if v.value is not None)
print(f"Extraction result reloaded: {populated}/{len(py_result.values)} values populated")
print("Comparison strategies:", [s.value for s in cmp_report.compared_strategies])
print("Parsed document pages:", reloaded_doc.page_count)
print("Profile table pages:", reloaded_profile.detected_table_count)
"""
)

md("## 11. Markdown preview (extraction result + comparison report)")
code(
    """
ext_md = (py_run_dir / "extraction_result.pymupdf.md").read_text()
cmp_md = (cmp_run_dir / "comparison_report.md").read_text()
print("===== extraction_result.pymupdf.md (first 1800 chars) =====\\n")
print(ext_md[:1800])
print("\\n\\n===== comparison_report.md (first 1800 chars) =====\\n")
print(cmp_md[:1800])
"""
)

md(
    """
## 11b. Excel export — shareable dictionary & results for a product manager

The pipeline already wrote `.xlsx` workbooks alongside the JSON/Markdown. Here
we confirm them and also show the standalone `ExcelExporter` API for ad-hoc
exports (e.g. emailing the dictionary to a PM).
"""
)
code(
    """
from large_pdf_extractor.rendering.excel_writer import ExcelExporter
import openpyxl

# Workbooks written automatically by the run:
dict_xlsx = py_run_dir / "extraction_dictionary.used.xlsx"
result_xlsx = py_run_dir / "extraction_result.pymupdf.xlsx"
print("Dictionary workbook:", dict_xlsx.exists(), dict_xlsx.name)
print("Results workbook:   ", result_xlsx.exists(), result_xlsx.name)

wb = openpyxl.load_workbook(dict_xlsx)
print("Dictionary sheets:", wb.sheetnames)
ws = wb["Dictionary"]
print("Columns:", [c.value for c in ws[1]])
print("Item rows:", ws.max_row - 1)

# Ad-hoc export to any path (e.g. to share directly):
pm_path = ExcelExporter().export_dictionary(validated, str(OUTPUT_DIR / "dictionary_for_PM.xlsx"))
print("\\nStandalone export for sharing:", pm_path)
"""
)

md("## 12. Artifact summary table")
code(
    """
import pandas as pd

rows = []
for ref in cmp_state.artifacts:
    rows.append({
        "artifact_type": ref.artifact_type.value,
        "strategy": ref.strategy.value if ref.strategy else "-",
        "file": Path(ref.path).name,
    })
summary_df = pd.DataFrame(rows)
summary_df
"""
)

md("## 13. Final acceptance checklist")
code(
    """
checks = {
    "Input PDF present": pdf_exists,
    "Profile computed (pages > 0)": profile.page_count > 0,
    "Dictionary proposed before extraction": len(dictionary.items) > 0,
    "Dictionary validated by Pydantic": isinstance(validated, ExtractionDictionary),
    "PyMuPDF extraction_result.json written": (py_run_dir / "extraction_result.pymupdf.json").exists(),
    "PyMuPDF extraction_result.md written": (py_run_dir / "extraction_result.pymupdf.md").exists(),
    "Docling path ran or skipped gracefully": (dl_run_dir / "parsed_document.docling.json").exists(),
    "Compare report JSON written": (cmp_run_dir / "comparison_report.json").exists(),
    "Compare report MD written": (cmp_run_dir / "comparison_report.md").exists(),
    "Shared dictionary used in compare": (cmp_run_dir / "extraction_dictionary.used.json").exists(),
    "Dictionary exported to Excel (.xlsx)": (py_run_dir / "extraction_dictionary.used.xlsx").exists(),
    "Outputs reload via Pydantic": isinstance(py_result, ExtractionResult),
    "At least one value extracted with source span": any(
        v.source_spans for v in py_result.values
    ),
}

all_pass = all(checks.values())
for name, ok in checks.items():
    print(f"[{'PASS' if ok else 'FAIL'}] {name}")
print("\\n" + ("ALL CHECKS PASSED ✅" if all_pass else "SOME CHECKS FAILED ❌"))
assert all_pass, "Phase 1 acceptance checklist failed."
"""
)

nb["cells"] = cells
out = Path(__file__).parent / "01_phase1_cme_reference_demo.ipynb"
nbf.write(nb, str(out))
print("Wrote", out)
