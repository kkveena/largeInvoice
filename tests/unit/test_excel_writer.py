"""Excel exporter tests."""

from __future__ import annotations

import openpyxl

from large_pdf_extractor.domain.models import ParseStrategy
from large_pdf_extractor.extraction.extractor import ExtractionEngine
from large_pdf_extractor.llm.fake_provider import FakeLLMProvider
from large_pdf_extractor.rendering.excel_writer import ExcelExporter


def test_export_dictionary_to_excel(tmp_path, sample_dictionary):
    out = tmp_path / "dict.xlsx"
    path = ExcelExporter().export_dictionary(sample_dictionary, str(out))
    assert out.exists()

    wb = openpyxl.load_workbook(path)
    assert "Dictionary" in wb.sheetnames
    assert "Overview" in wb.sheetnames

    ws = wb["Dictionary"]
    header = [c.value for c in ws[1]]
    assert "item_id" in header
    assert "instruction_prompt" in header
    # One header row + one row per dictionary item.
    assert ws.max_row == len(sample_dictionary.items) + 1


def test_export_extraction_result_to_excel(tmp_path, sample_chunks, sample_dictionary):
    result = ExtractionEngine(FakeLLMProvider()).extract(
        sample_dictionary, sample_chunks, ParseStrategy.PYMUPDF
    )
    out = tmp_path / "result.xlsx"
    path = ExcelExporter().export_extraction_result(result, str(out))
    assert out.exists()

    wb = openpyxl.load_workbook(path)
    assert "Extracted Values" in wb.sheetnames
    ws = wb["Extracted Values"]
    header = [c.value for c in ws[1]]
    assert "raw_value" in header
    assert "source_pages" in header
